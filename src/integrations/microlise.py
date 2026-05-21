"""Microlise TMS integration module.

Acts as the integration layer between the FPS back-office allocation system and the
Microlise Transport Management System (TMS). After the internal optimizer has produced a
vehicle-to-route allocation, this module translates FPS-internal vehicle IDs into Microlise
vehicle IDs and pushes each assignment to the Microlise Journeys Web API. It also generates
and uploads post-allocation Excel reports to Azure Blob Storage.

Environment variables consumed (all optional; empty string disables the feature):
  JLP_Microlise_TokenClientId          – OAuth2 client ID
  JLP_Microlise_TokenClient_Secret     – OAuth2 client secret
  JLP_Microlise_Token_URL              – Token endpoint URL
  JLP_Microlise_JourneysWebAPI_URL     – Journeys API base URL
  storage_account_conn_string          – Azure Storage connection string
  allocation_blob_container            – Blob container name
  allocation_blob_dir                  – Blob directory prefix
"""

import io
import os
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple

import requests
import requests.models
from requests.exceptions import (
    ConnectionError,
    HTTPError,
    RequestException,
    Timeout,
    TooManyRedirects,
)

from src.database.connection import db
from src.database.queries import Queries
from src.utils.logging_config import logger

# Alert message ID used for all Microlise API failures (matches t_alert.alert_message_id = 9)
_MICROLISE_ALERT_ID = 9


# --------------------------------------------------------------------------- #
# Parameter model                                                              #
# --------------------------------------------------------------------------- #


@dataclass
class MicroLiseParams:
    """Runtime parameters controlling a single Microlise dispatch run.

    Attributes:
        simulate_response: When True, skips real API calls and returns a synthetic 201.
        send_report: When True, generates and uploads the Excel allocation report.
        trigger_type: Allocation trigger type (e.g. 'initial', 'reallocation').
        initial_report: Include the morning/initial allocation sheet.
        compliance_report: Include the vehicle-match compliance sheet.
        unallocated_report: Include the unallocated routes sheet.
        start_hour_allocation: Upper bound (exclusive) of the qualifying initial-allocation
            window used when locating yesterday's compliance baseline.
        end_hour_allocation: Lower bound (inclusive) of that same window.
    """

    simulate_response: bool = True
    send_report: bool = False
    trigger_type: str = "initial"
    initial_report: bool = True
    compliance_report: bool = False
    unallocated_report: bool = False
    start_hour_allocation: int = 6
    end_hour_allocation: int = 4


# --------------------------------------------------------------------------- #
# Client                                                                       #
# --------------------------------------------------------------------------- #


class MicroLiseClient:
    """Integration client for the Microlise TMS Journeys Web API.

    Args:
        connection_type: ``'test'`` or ``'prod'``. Appended to alert messages when
            ``'test'`` so that on-call engineers can distinguish test traffic.
        client_id: OAuth2 client ID (falls back to ``JLP_Microlise_TokenClientId``).
        client_secret: OAuth2 client secret (falls back to env var).
        token_url: Token endpoint URL (falls back to env var).
        journeys_api_url: Journeys API base URL (falls back to env var).
        storage_conn_string: Azure Storage connection string (falls back to env var).
        blob_container: Blob container name (falls back to env var).
        blob_dir: Blob directory prefix (falls back to env var).
    """

    def __init__(
        self,
        connection_type: str = "prod",
        client_id: Optional[str] = None,
        client_secret: Optional[str] = None,
        token_url: Optional[str] = None,
        journeys_api_url: Optional[str] = None,
        storage_conn_string: Optional[str] = None,
        blob_container: Optional[str] = None,
        blob_dir: Optional[str] = None,
    ) -> None:
        self.connection_type = connection_type
        self.client_id = client_id or os.getenv("JLP_Microlise_TokenClientId", "")
        self.client_secret = client_secret or os.getenv("JLP_Microlise_TokenClient_Secret", "")
        self.token_url = token_url or os.getenv("JLP_Microlise_Token_URL", "")
        self.journeys_api_url = journeys_api_url or os.getenv("JLP_Microlise_JourneysWebAPI_URL", "")
        self.storage_conn_string = storage_conn_string or os.getenv("storage_account_conn_string", "")
        self.blob_container = blob_container or os.getenv("allocation_blob_container", "")
        self.blob_dir = blob_dir or os.getenv("allocation_blob_dir", "")

    # ----------------------------------------------------------------------- #
    # OAuth2 token                                                             #
    # ----------------------------------------------------------------------- #

    def get_token(self) -> str:
        """Fetch an OAuth2 Bearer token from the Microlise token endpoint.

        Raises:
            requests.exceptions.RequestException: On any network or HTTP failure.
        """
        logger.info("Fetching Microlise OAuth2 token")
        try:
            response = requests.post(
                self.token_url,
                auth=(self.client_id, self.client_secret),
                data={
                    "grant_type": "client_credentials",
                    "scope": "journeyallocatevehicle",
                },
                allow_redirects=False,
            )
            response.raise_for_status()
            token: str = response.json()["access_token"]
            logger.info("Successfully retrieved Microlise OAuth2 token")
            return token
        except (HTTPError, ConnectionError, Timeout, TooManyRedirects, RequestException) as exc:
            logger.error(f"Failed to retrieve Microlise OAuth2 token: {exc}")
            raise

    # ----------------------------------------------------------------------- #
    # Vehicle-ID translation                                                   #
    # ----------------------------------------------------------------------- #

    def get_vehicle_telematics_dict(self) -> Tuple[Dict[int, str], Dict[str, int]]:
        """Return bidirectional maps between FPS vehicle IDs and Microlise labels.

        Sentinel values ``{0: 'X', -1: '-1'}`` are always present.

        Returns:
            fps_to_microlise: ``{fps_vehicle_id -> microlise_label}``
            microlise_to_fps: ``{microlise_label -> fps_vehicle_id}``
        """
        if not db._connection or db._connection.closed:
            db.connect()
        
        rows = db.execute_query(Queries.GET_VEHICLE_TELEMATICS_DICT) or []
        fps_to_microlise: Dict[int, str] = {0: "X", -1: "-1"}
        microlise_to_fps: Dict[str, int] = {"X": 0, "-1": -1}
        for row in rows:
            vid = row["vehicle_id"]
            label = row["telematic_label"]
            fps_to_microlise[vid] = label
            microlise_to_fps[label] = vid
        logger.debug(f"Loaded {len(fps_to_microlise) - 2} vehicle telematics mappings")
        return fps_to_microlise, microlise_to_fps

    # ----------------------------------------------------------------------- #
    # API call & response handler                                              #
    # ----------------------------------------------------------------------- #

    def _post_vehicle_allocation(
        self, route_id: str, vehicle_name: str, token: str
    ) -> requests.models.Response:
        """POST a single vehicle-route assignment to the Microlise Journeys API."""
        url = f"{self.journeys_api_url}{route_id}/vehicles/"
        payload = {"VehicleName": vehicle_name, "scope": "journeyallocatevehicle"}
        headers = {
            "content-type": "text/json",
            "Authorization": f"Bearer {token}",
        }
        return requests.post(url, json=payload, headers=headers)

    def _build_simulated_response(self) -> requests.models.Response:
        """Return a synthetic 201 response without any network call."""
        resp = requests.models.Response()
        resp.status_code = 201
        return resp

    def http_response_handler(
        self,
        response: requests.models.Response,
        route_id: str,
        site_id: int,
    ) -> bool:
        """Handle the Microlise API response for a single route allocation.

        Inserts an alert row into ``t_alert`` for any non-success status code.

        Args:
            response: The HTTP response (real or simulated).
            route_id: Route identifier (for logging and alert context).
            site_id: Site identifier written to ``t_alert``.

        Returns:
            ``True`` for HTTP 200/201, ``False`` for all other codes.
        """
        code = response.status_code
        if code in (200, 201):
            logger.info(f"Microlise allocation success for route {route_id}: HTTP {code}")
            return True

        test_suffix = ": TEST" if self.connection_type == "test" else ""
        dev_app_id = (
            f"Vehicle allocation: Microlise server response - "
            f"{code} {response.text}{test_suffix}"
        )
        logger.warning(f"Microlise allocation failed for route {route_id}: {dev_app_id}")

        try:
            db.execute_query(
                Queries.INSERT_ALERT,
                (site_id, _MICROLISE_ALERT_ID, dev_app_id, datetime.now()),
                fetch=False,
            )
        except Exception as exc:
            logger.error(f"Failed to insert Microlise alert for route {route_id}: {exc}")

        return False

    # ----------------------------------------------------------------------- #
    # Main dispatch loop                                                       #
    # ----------------------------------------------------------------------- #

    def dispatch_allocations(
        self,
        allocation_id: int,
        site_id: int,
        params: MicroLiseParams,
    ) -> Tuple[Dict[str, Any], List[str]]:
        """Dispatch all pending vehicle-route assignments to the Microlise Journeys API.

        Fetches every ``t_route_allocated`` row for *site_id* whose ``status = 'N'``
        and ``http_response`` is not already in ``{200, 201, -1}``, then calls the
        Journeys API once per row (or synthesises a 201 in simulation mode). The HTTP
        response code is written back to ``t_route_allocated`` after each call.

        Finally the ``t_allocation_monitor.status`` column is set to ``'S'`` if every
        call succeeded or ``'F'`` if any failed.

        Args:
            allocation_id: FPS allocation run identifier.
            site_id: Site to scope route lookups.
            params: Runtime configuration for this dispatch run.

        Returns:
            Tuple of (result_dict, route_aliases) where *result_dict* contains
            dispatch statistics and *route_aliases* is the list of route aliases
            present in the current allocation (used for the unallocated report).
        """
        if not db._connection or db._connection.closed:
            db.connect()
        
        token: Optional[str] = None
        if not params.simulate_response:
            token = self.get_token()

        logger.info("Getting vehicle telematics dict")
        fps_to_microlise, _ = self.get_vehicle_telematics_dict()

        rows = db.execute_query(Queries.GET_ROUTES_FOR_DISPATCH) or []
        logger.info(f"Loaded {len(rows)} routes for dispatch")

        # print(f"Rows: {rows}")
        print(f"Site ID: {site_id}")
        print(f"HTTP response: {[r['http_response'] not in (200, 201, -1) for r in rows]}")

        pending = [
            r for r in rows
            if r["site_id"] == site_id
            and r["http_response"] not in (200, 201) # -1 is not dispatched yet
        ]
        
        print(f"Pending routes: {pending}")

        route_aliases: List[str] = [
            str(r.get("route_alias", "")) for r in rows if r["site_id"] == site_id
        ]

        print(f"Route aliases: {route_aliases}")

        logger.info(
            f"Dispatching {len(pending)} route(s) to Microlise "
            f"(allocation_id={allocation_id}, simulate={params.simulate_response})"
        )

        success_count = 0
        fail_count = 0

        for row in pending:
            vehicle_id: int = row["vehicle_id_allocated"]
            route_id = str(row["route_id"])
            microlise_label = fps_to_microlise.get(vehicle_id, "X")

            if params.simulate_response:
                response = self._build_simulated_response()
            else:
                response = self._post_vehicle_allocation(route_id, microlise_label, token)

            try:
                db.execute_query(
                    Queries.UPDATE_ROUTE_ALLOCATED_HTTP_RESPONSE,
                    (response.status_code, route_id),
                    fetch=False,
                )
            except Exception as exc:
                logger.error(f"Failed to persist http_response for route {route_id}: {exc}")

            if self.http_response_handler(response, route_id, site_id):
                success_count += 1
            else:
                fail_count += 1

        final_status = "S" if fail_count == 0 else "F"
        try:
            db.execute_query(
                Queries.UPDATE_ALLOCATION_MONITOR_STATUS,
                (final_status, allocation_id),
                fetch=False,
            )
        except Exception as exc:
            logger.error(f"Failed to update allocation monitor status: {exc}")

        logger.info(
            f"Microlise dispatch complete: {success_count} succeeded, "
            f"{fail_count} failed, monitor status={final_status}"
        )

        result: Dict[str, Any] = {
            "success": fail_count == 0,
            "routes_dispatched": success_count,
            "routes_failed": fail_count,
            "simulate_mode": params.simulate_response,
        }
        return result, route_aliases

    # ----------------------------------------------------------------------- #
    # Report generation                                                        #
    # ----------------------------------------------------------------------- #

    def find_missing_routes(
        self,
        site_id: int,
        allocation_route_aliases: List[str],
    ) -> List[str]:
        """Return route aliases that had Microlise fetch errors and are absent from the allocation.

        Queries ``t_error_log`` for today, filters to rows emitted by
        ``microlise-route-fetch`` containing the canonical route-alias error marker,
        then returns those route numbers not already present in
        *allocation_route_aliases*.
        """
        if not db._connection or db._connection.closed:
            db.connect()
        
        today_local = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        rows = (
            db.execute_query(
                Queries.GET_ERROR_LOG_FOR_DATE,
                (today_local, today_local + timedelta(days=1)),
            )
            or []
        )

        alias_set = set(allocation_route_aliases)
        missing: List[str] = []
        marker = "Issues with TMC API route alias call for route num "

        for row in rows:
            module_no = str(row.get("module_no", ""))
            message = str(row.get("error_message", ""))
            if "microlise-route-fetch" not in module_no:
                continue
            if marker not in message:
                continue
            route_num = message.split(marker)[-1].strip()
            if route_num not in alias_set:
                missing.append(route_num)

        return missing

    def _generate_compliance_sheet(
        self,
        site_id: int,
        fps_to_microlise: Dict[int, str],
        params: MicroLiseParams,
    ) -> Optional[Any]:
        """Build the compliance DataFrame for yesterday's initial allocation, or None.

        Locates an initial allocation row in ``t_allocation_monitor`` from the
        previous day whose ``run_datetime`` hour falls within the configured
        allocation window (``end_hour_allocation <= hour < start_hour_allocation``).
        Merges the corresponding history against ``t_route_plan`` to produce a
        Vehicle Match comparison, retaining only mismatched rows.
        """
        try:
            import pandas as pd
        except ImportError:
            logger.error("pandas not installed; cannot generate compliance report")
            return None

        if not db._connection or db._connection.closed:
            db.connect()

        report_date = (datetime.now() - timedelta(days=1)).date()
        rows = (
            db.execute_query(Queries.GET_ALLOCATION_MONITOR_BY_DATE, (report_date,)) or []
        )

        qualifying = [
            r
            for r in rows
            if r.get("trigger_type") == "initial"
            and params.end_hour_allocation
            <= r["run_datetime"].hour
            < params.start_hour_allocation
        ]
        if len(qualifying) != 1:
            logger.info(
                f"Compliance report: expected 1 qualifying allocation for {report_date}, "
                f"found {len(qualifying)} – skipping"
            )
            return None

        hist_allocation_id = qualifying[0]["allocation_id"]
        hist_rows = (
            db.execute_query(
                Queries.GET_ROUTE_ALLOCATED_HISTORY_BY_IDS, (hist_allocation_id,)
            )
            or []
        )
        if not hist_rows:
            return None

        route_ids = tuple(r["route_id"] for r in hist_rows)
        plan_rows = db.execute_query(Queries.GET_ROUTE_PLAN_BY_IDS, (route_ids,)) or []

        hist_df = pd.DataFrame(hist_rows)
        plan_df = pd.DataFrame(plan_rows)
        if hist_df.empty or plan_df.empty:
            return None

        merged = hist_df.merge(
            plan_df[["route_id", "vehicle_id"]],
            on="route_id",
            how="left",
            suffixes=("_hist", "_plan"),
        )
        merged["vehicle_id_allocated_label"] = merged["vehicle_id_allocated"].map(
            fps_to_microlise
        )
        merged["Vehicle Match"] = (
            merged["vehicle_id_allocated"] == merged["vehicle_id_plan"]
        )

        route_status_col = merged.get("route_status", pd.Series(dtype=str))
        compliance = merged.loc[
            ~merged["Vehicle Match"]
            & ~route_status_col.isin(["X", "E"])
        ][
            [
                "route_id",
                "route_alias",
                "vehicle_id_allocated_label",
                "vehicle_id_plan",
                "Vehicle Match",
            ]
        ].copy()

        compliance.columns = [
            "Route ID",
            "Route Number",
            "Vehicle ID Allocated",
            "Vehicle ID Used",
            "Vehicle Match",
        ]
        return compliance if not compliance.empty else None

    def save_allocation_report(
        self,
        allocation_id: int,
        site_id: int,
        fps_to_microlise: Dict[int, str],
        params: MicroLiseParams,
        route_aliases: List[str],
    ) -> None:
        """Assemble a multi-sheet Excel workbook and upload it to Azure Blob Storage.

        Sheets included are controlled by the flags on *params*:
          - ``Morning Report YYYY_MM_DD``   (params.initial_report)
          - ``Compliance Report YYYY_MM_DD`` (params.compliance_report)
          - ``Unallocated Report YYYY_MM_DD``(params.unallocated_report)

        Only runs when ``params.send_report`` is ``True``, ``trigger_type`` is
        ``'initial'``, and ``simulate_response`` is ``False``.
        """
        try:
            from openpyxl import Workbook
        except ImportError:
            logger.error("openpyxl not installed; cannot generate allocation report")
            return

        if not db._connection or db._connection.closed:
            db.connect()

        today_str = datetime.now().strftime("%Y_%m_%d")
        wb = Workbook()
        wb.remove(wb.active)
        sheets_added = 0

        if params.initial_report:
            rows = db.execute_query(Queries.GET_ROUTES_FOR_DISPATCH) or []
            site_rows = [r for r in rows if r["site_id"] == site_id]
            ws = wb.create_sheet(title=f"Morning Report {today_str}")
            ws.append(["Route ID", "Vehicle ID", "Route Number"])
            for row in site_rows:
                label = fps_to_microlise.get(row["vehicle_id_allocated"], "X")
                ws.append([str(row["route_id"]), label, str(row.get("route_alias", ""))])
            sheets_added += 1

        if params.compliance_report:
            compliance_df = self._generate_compliance_sheet(
                site_id, fps_to_microlise, params
            )
            if compliance_df is not None:
                ws = wb.create_sheet(title=f"Compliance Report {today_str}")
                ws.append(list(compliance_df.columns))
                for _, r in compliance_df.iterrows():
                    ws.append(list(r))
                sheets_added += 1

        if params.unallocated_report:
            missing = self.find_missing_routes(site_id, route_aliases)
            if missing:
                ws = wb.create_sheet(title=f"Unallocated Report {today_str}")
                ws.append(["Unallocated Route"])
                for alias in missing:
                    ws.append([alias])
                sheets_added += 1

        if sheets_added == 0:
            logger.info("No report sheets produced; skipping blob upload")
            return

        report_file_name = f"daily_allocation_report_{site_id}_{today_str}"
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        self._upload_blob(buffer, report_file_name)

    def _upload_blob(self, buffer: io.BytesIO, report_file_name: str) -> None:
        """Upload an in-memory workbook to Azure Blob Storage.

        Any exception during upload is caught and logged; it does not propagate.
        """
        try:
            from azure.storage.blob import BlobClient, ContentSettings
        except ImportError:
            logger.error(
                "azure-storage-blob not installed; "
                "install it with: pip install azure-storage-blob"
            )
            return

        blob_name = f"{self.blob_dir}/{report_file_name}.xlsx"
        try:
            blob_client = BlobClient.from_connection_string(
                self.storage_conn_string,
                self.blob_container,
                blob_name,
            )
            blob_client.upload_blob(
                buffer,
                overwrite=True,
                content_settings=ContentSettings(content_type="file/xlsx"),
            )
            logger.info(f"Allocation report uploaded to Azure Blob: {blob_name}")
        except Exception as exc:
            logger.error(f"Failed to upload report to Azure Blob Storage: {exc}")

    # ----------------------------------------------------------------------- #
    # Main entry point                                                         #
    # ----------------------------------------------------------------------- #

    def run(
        self,
        allocation_id: int,
        site_id: int,
        params: MicroLiseParams,
    ) -> Dict[str, Any]:
        """Execute the full Microlise post-allocation workflow.

        1. Dispatches vehicle-route assignments to the Journeys API (or simulates).
        2. Optionally generates and uploads the Excel allocation report to blob storage.

        Report upload only occurs when all three conditions are met:
          - ``params.send_report`` is ``True``
          - ``params.trigger_type == 'initial'``
          - ``params.simulate_response`` is ``False``

        Args:
            allocation_id: FPS allocation run ID (from ``t_allocation_monitor``).
            site_id: Site identifier.
            params: Runtime parameters.

        Returns:
            Dict suitable for inclusion in the API response.
        """
        logger.info(
            f"Starting Microlise integration run: allocation_id={allocation_id}, "
            f"site_id={site_id}, simulate={params.simulate_response}, "
            f"trigger_type={params.trigger_type}"
        )

        dispatch_result, route_aliases = self.dispatch_allocations(
            allocation_id, site_id, params
        )

        report_uploaded = False
        if (
            params.send_report
            and params.trigger_type == "initial"
            and not params.simulate_response
            and (params.initial_report or params.compliance_report or params.unallocated_report)
        ):
            fps_to_microlise, _ = self.get_vehicle_telematics_dict()
            self.save_allocation_report(
                allocation_id,
                site_id,
                fps_to_microlise,
                params,
                route_aliases,
            )
            report_uploaded = True

        return {
            **dispatch_result,
            "allocation_id": allocation_id,
            "report_uploaded": report_uploaded,
        }
